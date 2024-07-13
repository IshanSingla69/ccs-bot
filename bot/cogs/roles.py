import discord
from discord.ext import commands
import os
import openpyxl
import logging
import asyncio

logging.basicConfig(
    filename="bot.log",
    level=logging.INFO,
    format="%(asctime)s:%(levelname)s:%(message)s",
)


class Roles(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    @commands.command(
        name="export_roles_xlsx", help="Export roles to an Excel file", category="Roles"
    )
    @commands.has_any_role(768960824009162802, 1254871511056257144)
    async def export_roles(self, ctx):
        guild = ctx.guild
        if guild:
            logging.info("Starting export_roles_xlsx command execution.")
            filename = "roles.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Roles"

            # Add headers
            logging.info(f"Generating file: {filename}")
            ws.append(["Role Name", "Role ID"])

            # Iterate over roles and add them to the worksheet
            for role in guild.roles:
                ws.append([role.name, role.id])

            # Save the workbook
            wb.save(filename)
            logging.info(f"Sending file: {filename}")
            await ctx.send(file=discord.File(filename))
            os.remove(filename)
            logging.info("Completed export_roles_xlsx command execution.")
        else:
            await ctx.send("Guild not found!")

    @commands.command(name="delete_roles_xlsx")
    @commands.has_any_role(768960824009162802, 1254871511056257144)
    async def delete_roles_xlsx(self, ctx):
        await ctx.send("Please upload the Excel (.xlsx) file to delete roles.")

        def check(message):
            return (
                message.author == ctx.author
                and message.channel == ctx.channel
                and len(message.attachments) > 0
                and message.attachments[0].filename.endswith(".xlsx")
            )

        try:
            message = await self.bot.wait_for("message", check=check, timeout=60)
            await self.handle_excel_file(message)
        except asyncio.TimeoutError:
            await ctx.send("File upload timed out. Please try again.")

    async def handle_excel_file(self, message):
        # Get the attachment
        attachment = message.attachments[0]

        # Download the attachment and save with a unique name
        file_path = os.path.join(os.getcwd(), attachment.filename)
        await attachment.save(file_path)

        # Open the workbook
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Find the column index for 'Role ID'
            role_id_column = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "Role ID":
                    role_id_column = col
                    break

            if role_id_column is None:
                await message.channel.send("No 'Role ID' column found.")
                return

            guild = message.guild
            if guild:
                roles_to_delete = set()

                # Collect role IDs from the worksheet
                for row in range(2, ws.max_row + 1):
                    role_id = ws.cell(row=row, column=role_id_column).value
                    if role_id:
                        roles_to_delete.add(int(role_id))  # Convert to int if needed

                if not roles_to_delete:
                    await message.channel.send("No role IDs provided.")
                else:
                    # Delete roles based on provided role IDs
                    deleted_roles = []
                    for role_id in roles_to_delete:
                        role = guild.get_role(role_id)
                        if role:
                            try:
                                await role.delete()
                                deleted_roles.append(role.name)
                            except discord.Forbidden:
                                await message.channel.send(
                                    f"Failed to delete role {role.name}. Missing permissions."
                                )
                            except discord.HTTPException as e:
                                await message.channel.send(
                                    f"Failed to delete role {role.name}. HTTPException: {e}"
                                )

                    if deleted_roles:
                        await message.channel.send(
                            f"Deleted roles: {', '.join(deleted_roles)}"
                        )
                    else:
                        await message.channel.send(
                            "No roles deleted. Please check if the role IDs are correct."
                        )

            else:
                await message.channel.send("Guild not found.")

        except Exception as e:
            await message.channel.send(f"Error processing Excel file: {str(e)}")

        finally:
            # Delete the file after processing
            os.remove(file_path)


async def setup(bot):
    await bot.add_cog(Roles(bot))
